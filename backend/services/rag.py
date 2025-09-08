from __future__ import annotations

import hashlib
import json
import os
from typing import Any, Dict, List, Optional

import faiss
import numpy as np
from flask import current_app


DIM = 128


def _index_paths() -> tuple[str, str]:
    base = os.path.join(current_app.root_path, "rag_index")
    os.makedirs(base, exist_ok=True)
    return os.path.join(base, "index.faiss"), os.path.join(base, "meta.jsonl")


def _load_index() -> tuple[faiss.IndexFlatL2, List[Dict[str, Any]]]:
    idx_path, meta_path = _index_paths()
    if os.path.exists(idx_path):
        index = faiss.read_index(idx_path)
    else:
        index = faiss.IndexFlatL2(DIM)
    meta: List[Dict[str, Any]] = []
    if os.path.exists(meta_path):
        with open(meta_path, "r", encoding="utf-8") as f:
            meta = [json.loads(line) for line in f]
    return index, meta


def _save_index(index: faiss.IndexFlatL2, meta: List[Dict[str, Any]]) -> None:
    idx_path, meta_path = _index_paths()
    faiss.write_index(index, idx_path)
    with open(meta_path, "w", encoding="utf-8") as f:
        for m in meta:
            f.write(json.dumps(m, ensure_ascii=False) + "\n")


def _embed(text: str) -> np.ndarray:
    vec = np.zeros(DIM, dtype="float32")
    for token in text.lower().split():
        h = int(hashlib.md5(token.encode()).hexdigest(), 16)
        vec[h % DIM] += 1.0
    norm = np.linalg.norm(vec)
    if norm > 0:
        vec /= norm
    return vec


def _extract_attachment_text(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext in {".txt", ".csv"}:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        if ext in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook

            wb = load_workbook(path, data_only=True)
            text: List[str] = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            text.append(str(cell))
            return "\n".join(text)
        if ext == ".pdf":
            try:
                import PyPDF2

                text = []
                with open(path, "rb") as f:
                    reader = PyPDF2.PdfReader(f)
                    for page in reader.pages:
                        text.append(page.extract_text() or "")
                return "\n".join(text)
            except Exception:
                pass
        if ext in {".png", ".jpg", ".jpeg"}:
            try:
                from PIL import Image
                import pytesseract

                img = Image.open(path)
                return pytesseract.image_to_string(img)
            except Exception:
                pass
    except Exception:
        pass
    return ""


def index_response(resp_text: str, meta_info: Dict[str, Any], attachments: Optional[List[str]] = None) -> None:
    index, meta = _load_index()
    full_text = resp_text
    if attachments:
        for p in attachments:
            extracted = _extract_attachment_text(p)
            if extracted:
                full_text += "\n" + extracted
    chunks = [full_text[i : i + 200] for i in range(0, len(full_text), 200)]
    for chunk in chunks:
        vec = _embed(chunk)
        index.add(np.expand_dims(vec, axis=0))
        meta.append({"text": chunk, **meta_info})
    _save_index(index, meta)


def search(query: str, k: int = 5) -> List[Dict[str, Any]]:
    index, meta = _load_index()
    if index.ntotal == 0:
        return []
    vec = _embed(query)
    D, I = index.search(np.expand_dims(vec, axis=0), k)
    results: List[Dict[str, Any]] = []
    for idx in I[0]:
        if 0 <= idx < len(meta):
            results.append(meta[idx])
    return results
