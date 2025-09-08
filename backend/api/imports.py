from __future__ import annotations

import csv
from datetime import date, datetime
from io import TextIOWrapper, BytesIO

from flask import Response, jsonify, request
from openpyxl import Workbook, load_workbook

from . import api_bp
from ..app import db
from ..models import (
    AuditLog,
    BudgetAnnual,
    BudgetSnapshot,
    OpsActualDaily,
    CalendarDim,
    SeasonalityMonth,
    RecomputeJob,
    Service,
    Unit,
)


class RowError(dict):
    row: int
    errors: list[str]


def _parse_csv(file, required: list[str]) -> tuple[list[dict[str, str]], list[RowError]]:
    stream = TextIOWrapper(file.stream, encoding="utf-8")
    reader = csv.DictReader(stream)
    rows: list[dict[str, str]] = []
    errors: list[RowError] = []
    for idx, row in enumerate(reader, start=2):
        row_errors = [f"missing {col}" for col in required if not row.get(col)]
        if row_errors:
            errors.append({"row": idx, "errors": row_errors})
        rows.append(row)
    return rows, errors


def _parse_services(file) -> tuple[list[dict[str, str]], list[RowError]]:
    """Parse services from CSV or XLSX and validate fields."""
    required = ["code", "name", "uom", "is_active"]
    rows: list[dict[str, str]] = []
    errors: list[RowError] = []
    codes: set[str] = set()
    if file.filename and file.filename.lower().endswith(".xlsx"):
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        for idx, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = {
                "code": r[0] if r and len(r) > 0 and r[0] is not None else "",
                "name": r[1] if r and len(r) > 1 and r[1] is not None else "",
                "uom": r[2] if r and len(r) > 2 and r[2] is not None else "",
                "base_qty": r[3] if r and len(r) > 3 else None,
                "base_fin": r[4] if r and len(r) > 4 else None,
                "is_active": r[5] if r and len(r) > 5 and r[5] is not None else "",
            }
            row_errors: list[str] = [f"missing {col}" for col in required if not data.get(col)]
            code = str(data.get("code"))
            if code in codes:
                row_errors.append("duplicate code")
            else:
                codes.add(code)
            ia = str(data.get("is_active"))
            if ia not in ("0", "1"):
                row_errors.append("is_active must be 0 or 1")
            for field in ["base_qty", "base_fin"]:
                if data.get(field) not in (None, ""):
                    try:
                        float(data[field])
                    except ValueError:
                        row_errors.append(f"{field} must be number")
            if row_errors:
                errors.append({"row": idx, "errors": row_errors})
            rows.append({k: ("" if v is None else str(v)) for k, v in data.items()})
    else:
        rows, errors = _parse_csv(file, required)
        codes = set()
        for idx, row in enumerate(rows, start=2):
            row_errors: list[str] = []
            code = row.get("code", "")
            if code in codes:
                row_errors.append("duplicate code")
            else:
                codes.add(code)
            ia = row.get("is_active")
            if ia not in ("0", "1"):
                row_errors.append("is_active must be 0 or 1")
            for field in ["base_qty", "base_fin"]:
                if row.get(field):
                    try:
                        float(row[field])
                    except ValueError:
                        row_errors.append(f"{field} must be number")
            if row_errors:
                errors.append({"row": idx, "errors": row_errors})
    return rows, errors


def _parse_budget(file) -> tuple[list[dict[str, str]], list[RowError]]:
    """Parse BudgetAnnual from CSV or XLSX and validate fields."""
    required = ["year", "scenario", "version", "service_code", "currency"]
    rows: list[dict[str, str]] = []
    errors: list[RowError] = []
    if file.filename and file.filename.lower().endswith(".xlsx"):
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        for idx, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = {
                "year": r[0] if r and len(r) > 0 and r[0] is not None else "",
                "scenario": r[1] if r and len(r) > 1 and r[1] is not None else "",
                "version": r[2] if r and len(r) > 2 and r[2] is not None else "",
                "service_code": r[3] if r and len(r) > 3 and r[3] is not None else "",
                "unit_id": r[4] if r and len(r) > 4 else None,
                "annual_qty": r[5] if r and len(r) > 5 else None,
                "annual_fin": r[6] if r and len(r) > 6 else None,
                "currency": r[7] if r and len(r) > 7 and r[7] is not None else "",
                "notes": r[8] if r and len(r) > 8 else None,
            }
            row_errors = [f"missing {col}" for col in required if not data.get(col)]
            for field in ["year", "unit_id"]:
                val = data.get(field)
                if val not in (None, ""):
                    try:
                        int(val)
                    except ValueError:
                        row_errors.append(f"{field} must be int")
            for field in ["annual_qty", "annual_fin"]:
                val = data.get(field)
                if val not in (None, ""):
                    try:
                        float(val)
                    except ValueError:
                        row_errors.append(f"{field} must be number")
            if row_errors:
                errors.append({"row": idx, "errors": row_errors})
            rows.append({k: ("" if v is None else str(v)) for k, v in data.items()})
    else:
        rows, errors = _parse_csv(file, required)
        for idx, row in enumerate(rows, start=2):
            row_errors: list[str] = []
            for field in ["year", "unit_id"]:
                if row.get(field):
                    try:
                        int(row[field])
                    except ValueError:
                        row_errors.append(f"{field} must be int")
            for field in ["annual_qty", "annual_fin"]:
                if row.get(field):
                    try:
                        float(row[field])
                    except ValueError:
                        row_errors.append(f"{field} must be number")
            if row_errors:
                errors.append({"row": idx, "errors": row_errors})
    return rows, errors


def _parse_ops_actual(file) -> tuple[list[dict[str, str]], list[RowError]]:
    """Parse OpsActualDaily from CSV or XLSX and validate basic fields."""
    required = ["date_shamsi", "service_code", "unit_id"]
    rows: list[dict[str, str]] = []
    errors: list[RowError] = []
    if file.filename and file.filename.lower().endswith(".xlsx"):
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        for idx, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = {
                "date_shamsi": r[0] if r and len(r) > 0 and r[0] is not None else "",
                "date_gregorian": r[1] if r and len(r) > 1 and r[1] is not None else "",
                "service_code": r[2] if r and len(r) > 2 and r[2] is not None else "",
                "unit_id": r[3] if r and len(r) > 3 else None,
                "qty": r[4] if r and len(r) > 4 else None,
                "fin": r[5] if r and len(r) > 5 else None,
            }
            row_errors = [f"missing {col}" for col in required if not data.get(col)]
            # validate date
            ds = data.get("date_shamsi")
            if ds:
                try:
                    date.fromisoformat(str(ds).replace("/", "-"))
                except Exception:
                    row_errors.append("invalid date_shamsi")
            # unit id must be int
            if data.get("unit_id") not in (None, ""):
                try:
                    int(data["unit_id"])
                except ValueError:
                    row_errors.append("unit_id must be int")
            # qty/fin must be non-negative numbers
            for field in ["qty", "fin"]:
                val = data.get(field)
                if val not in (None, ""):
                    try:
                        if float(val) < 0:
                            row_errors.append(f"{field} must be non-negative")
                    except ValueError:
                        row_errors.append(f"{field} must be number")
            if row_errors:
                errors.append({"row": idx, "errors": row_errors})
            rows.append({k: ("" if v is None else str(v)) for k, v in data.items()})
    else:
        rows, errors = _parse_csv(file, required)
        for idx, row in enumerate(rows, start=2):
            row_errors: list[str] = []
            ds = row.get("date_shamsi")
            if ds:
                try:
                    date.fromisoformat(ds.replace("/", "-"))
                except Exception:
                    row_errors.append("invalid date_shamsi")
            if row.get("unit_id"):
                try:
                    int(row["unit_id"])
                except ValueError:
                    row_errors.append("unit_id must be int")
            for field in ["qty", "fin"]:
                if row.get(field):
                    try:
                        if float(row[field]) < 0:
                            row_errors.append(f"{field} must be non-negative")
                    except ValueError:
                        row_errors.append(f"{field} must be number")
            if row_errors:
                errors.append({"row": idx, "errors": row_errors})
    return rows, errors


def _parse_calendar(file) -> tuple[list[dict[str, str]], list[RowError]]:
    required = [
        "date_shamsi",
        "jalali_month",
        "weekday_name",
        "is_friday",
        "is_thursday",
        "is_official_holiday",
        "is_summer_break",
    ]
    rows: list[dict[str, str]] = []
    errors: list[RowError] = []
    if file.filename and file.filename.lower().endswith(".xlsx"):
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        for idx, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = {
                "date_shamsi": r[0] if r and len(r) > 0 and r[0] is not None else "",
                "jalali_month": r[1] if r and len(r) > 1 and r[1] is not None else "",
                "weekday_name": r[2] if r and len(r) > 2 and r[2] is not None else "",
                "is_friday": r[3] if r and len(r) > 3 and r[3] is not None else "",
                "is_thursday": r[4] if r and len(r) > 4 and r[4] is not None else "",
                "is_official_holiday": r[5] if r and len(r) > 5 and r[5] is not None else "",
                "is_summer_break": r[6] if r and len(r) > 6 and r[6] is not None else "",
            }
            row_errors: list[str] = [f"missing {col}" for col in required if not data.get(col)]
            for field in ["jalali_month"]:
                val = data.get(field)
                if val not in (None, ""):
                    try:
                        int(val)
                    except ValueError:
                        row_errors.append(f"{field} must be int")
            for field in [
                "is_friday",
                "is_thursday",
                "is_official_holiday",
                "is_summer_break",
            ]:
                val = str(data.get(field))
                if val not in ("0", "1"):
                    row_errors.append(f"{field} must be 0 or 1")
            weight = 1.0
            if data.get("is_friday") == "1" or data.get("is_official_holiday") == "1" or data.get("is_summer_break") == "1":
                weight = 0.0
            elif data.get("is_thursday") == "1":
                weight = 0.5
            data["weight_raw"] = str(weight)
            if row_errors:
                errors.append({"row": idx, "errors": row_errors})
            rows.append(data)
    else:
        rows, errors = _parse_csv(file, required)
        for idx, row in enumerate(rows, start=2):
            row_errors: list[str] = []
            for field in ["jalali_month"]:
                if row.get(field):
                    try:
                        int(row[field])
                    except ValueError:
                        row_errors.append(f"{field} must be int")
            for field in [
                "is_friday",
                "is_thursday",
                "is_official_holiday",
                "is_summer_break",
            ]:
                val = row.get(field)
                if val not in ("0", "1"):
                    row_errors.append(f"{field} must be 0 or 1")
            weight = 1.0
            if row.get("is_friday") == "1" or row.get("is_official_holiday") == "1" or row.get("is_summer_break") == "1":
                weight = 0.0
            elif row.get("is_thursday") == "1":
                weight = 0.5
            row["weight_raw"] = str(weight)
            if row_errors:
                errors.append({"row": idx, "errors": row_errors})
    return rows, errors


def _parse_seasonality(file) -> tuple[list[dict[str, str]], list[RowError]]:
    required = ["month", "actual_1403"]
    rows: list[dict[str, str]] = []
    errors: list[RowError] = []
    if file.filename and file.filename.lower().endswith(".xlsx"):
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        for idx, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = {
                "month": r[0] if r and len(r) > 0 and r[0] is not None else "",
                "actual_1403": r[1] if r and len(r) > 1 and r[1] is not None else "",
            }
            row_errors: list[str] = [f"missing {col}" for col in required if not data.get(col)]
            if data.get("month") not in (None, ""):
                try:
                    m = int(data["month"])
                    if m < 1 or m > 12:
                        row_errors.append("month out of range")
                except ValueError:
                    row_errors.append("month must be int")
            if data.get("actual_1403") not in (None, ""):
                try:
                    float(data["actual_1403"])
                except ValueError:
                    row_errors.append("actual_1403 must be number")
            if row_errors:
                errors.append({"row": idx, "errors": row_errors})
            rows.append(data)
    else:
        rows, errors = _parse_csv(file, required)
        for idx, row in enumerate(rows, start=2):
            row_errors: list[str] = []
            if row.get("month"):
                try:
                    m = int(row["month"])
                    if m < 1 or m > 12:
                        row_errors.append("month out of range")
                except ValueError:
                    row_errors.append("month must be int")
            if row.get("actual_1403"):
                try:
                    float(row["actual_1403"])
                except ValueError:
                    row_errors.append("actual_1403 must be number")
            if row_errors:
                errors.append({"row": idx, "errors": row_errors})
    total = sum(float(r["actual_1403"]) for r in rows if r.get("actual_1403") not in (None, "")) or 1.0
    for r in rows:
        try:
            w = float(r["actual_1403"]) / total
        except Exception:
            w = 0.0
        r["season_weight"] = str(w)
    return rows, errors


@api_bp.post("/import/services")
def import_services() -> Response:
    mode = request.args.get("mode", "dry-run")
    file = request.files.get("file")
    if not file:
        return jsonify({"errors": [{"row": 0, "errors": ["no file"]}]}), 400
    rows, errors = _parse_services(file)
    existing_codes = {s.code for s in Service.query.all()}
    file_codes = {r["code"] for r in rows}
    created_list = sorted(list(file_codes - existing_codes))
    updated_list = sorted(list(file_codes & existing_codes))
    deactivated_list = sorted(list(existing_codes - file_codes))
    if mode == "dry-run":
        return jsonify(
            {
                "created": created_list,
                "updated": updated_list,
                "deactivated": deactivated_list,
                "errors": errors,
            }
        )
    if errors:
        return jsonify({"errors": errors}), 400
    confirm_deactivate = request.args.get("confirm_deactivate") == "true"
    created = updated = deactivated = 0
    for row in rows:
        code = row["code"]
        svc = Service.query.get(code)
        is_active = row.get("is_active") == "1"
        base_qty = float(row["base_qty"]) if row.get("base_qty") else None
        base_fin = float(row["base_fin"]) if row.get("base_fin") else None
        if svc:
            svc.name = row["name"]
            svc.uom = row["uom"]
            svc.base_qty = base_qty
            svc.base_fin = base_fin
            svc.is_active = is_active
            updated += 1
            action = "update"
        else:
            svc = Service(
                code=code,
                name=row["name"],
                uom=row["uom"],
                base_qty=base_qty,
                base_fin=base_fin,
                is_active=is_active,
            )
            db.session.add(svc)
            created += 1
            action = "create"
        db.session.add(
            AuditLog(
                entity="service",
                entity_id=code,
                action=action,
                actor="system",
                at=datetime.utcnow(),
                diff_json={},
            )
        )
    if confirm_deactivate:
        for code in deactivated_list:
            svc = Service.query.get(code)
            if svc and svc.is_active:
                svc.is_active = False
                deactivated += 1
                db.session.add(
                    AuditLog(
                        entity="service",
                        entity_id=code,
                        action="deactivate",
                        actor="system",
                        at=datetime.utcnow(),
                        diff_json={},
                    )
                )
    db.session.commit()
    return jsonify(
        {"created": created, "updated": updated, "deactivated": deactivated}
    )


@api_bp.post("/import/budget-annual")
def import_budget() -> Response:
    mode = request.args.get("mode", "dry-run")
    file = request.files.get("file")
    if not file:
        return jsonify({"errors": [{"row": 0, "errors": ["no file"]}]}), 400
    rows, errors = _parse_budget(file)
    created_list: list[str] = []
    updated_list: list[str] = []
    for row in rows:
        key = dict(
            year=int(row["year"]),
            scenario=row["scenario"],
            version=row["version"],
            service_code=row["service_code"],
            unit_id=int(row["unit_id"]) if row.get("unit_id") else None,
        )
        if BudgetAnnual.query.filter_by(**key).first():
            updated_list.append(row["service_code"])
        else:
            created_list.append(row["service_code"])
    if mode == "dry-run":
        return jsonify(
            {
                "errors": errors,
                "created": created_list,
                "updated": updated_list,
                "deactivated": [],
            }
        )
    if errors:
        return jsonify({"errors": errors}), 400
    snapshots: dict[tuple[int, str], int] = {}
    # create snapshot per (year, scenario)
    for y, sc in {(int(r["year"]), r["scenario"]) for r in rows}:
        BudgetSnapshot.query.filter_by(year=y, scenario=sc, status="published").update(
            {"status": "archived"}
        )
        snap = BudgetSnapshot(year=y, scenario=sc, status="published")
        db.session.add(snap)
        db.session.flush()
        snapshots[(y, sc)] = snap.id
    created = updated = 0
    for row in rows:
        y = int(row["year"])
        sc = row["scenario"]
        key = dict(
            year=y,
            scenario=sc,
            version=row["version"],
            service_code=row["service_code"],
            unit_id=int(row["unit_id"]) if row.get("unit_id") else None,
        )
        entry = BudgetAnnual.query.filter_by(**key).first()
        annual_qty = float(row["annual_qty"]) if row.get("annual_qty") else None
        annual_fin = float(row["annual_fin"]) if row.get("annual_fin") else None
        notes = row.get("notes") or None
        if entry:
            entry.annual_qty = annual_qty
            entry.annual_fin = annual_fin
            entry.currency = row["currency"]
            entry.snapshot_id = snapshots[(y, sc)]
            entry.notes = notes
            updated += 1
            action = "update"
        else:
            entry = BudgetAnnual(
                **key,
                annual_qty=annual_qty,
                annual_fin=annual_fin,
                currency=row["currency"],
                snapshot_id=snapshots[(y, sc)],
                notes=notes,
            )
            db.session.add(entry)
            created += 1
            action = "create"
        db.session.add(
            AuditLog(
                entity="budget_annual",
                entity_id=str(entry.id if entry.id else 0),
                action=action,
                actor="system",
                at=datetime.utcnow(),
                diff_json={},
            )
        )
    db.session.add(RecomputeJob(job="budget"))
    db.session.commit()
    from ..services.recompute import run_recompute_jobs

    run_recompute_jobs(db.session)
    return jsonify({"created": created, "updated": updated, "deactivated": 0})


@api_bp.post("/import/ops-actual")
def import_ops_actual() -> Response:
    mode = request.args.get("mode", "dry-run")
    file = request.files.get("file")
    if not file:
        return jsonify({"errors": [{"row": 0, "errors": ["no file"]}]}), 400
    rows, errors = _parse_ops_actual(file)
    # validate service and unit existence and gather created/updated keys
    services = {s.code for s in Service.query.all()}
    units = {u.id for u in Unit.query.all()}
    existing = {
        (o.date.isoformat(), o.service_code, o.unit_id)
        for o in OpsActualDaily.query.all()
    }
    created_keys: list[str] = []
    updated_keys: list[str] = []
    for idx, row in enumerate(rows, start=2):
        row_errors: list[str] = []
        if row.get("service_code") not in services:
            row_errors.append("unknown service_code")
        try:
            uid = int(row.get("unit_id", "0"))
            if uid not in units:
                row_errors.append("unknown unit_id")
        except ValueError:
            pass
        if row_errors:
            errors.append({"row": idx, "errors": row_errors})
        key = (
            row.get("date_shamsi", "").replace("/", "-"),
            row.get("service_code"),
            int(row.get("unit_id", "0") or 0),
        )
        if key in existing:
            updated_keys.append(f"{key[0]}-{key[1]}-{key[2]}")
        else:
            created_keys.append(f"{key[0]}-{key[1]}-{key[2]}")
    if mode == "dry-run":
        return jsonify({"errors": errors, "created": created_keys, "updated": updated_keys})
    if errors:
        return jsonify({"errors": errors}), 400
    created = updated = 0
    for row in rows:
        d = date.fromisoformat(row["date_shamsi"].replace("/", "-"))
        sc = row["service_code"]
        uid = int(row["unit_id"])
        qty = float(row["qty"]) if row.get("qty") else None
        fin = float(row["fin"]) if row.get("fin") else None
        rec = OpsActualDaily.query.filter_by(
            date=d, service_code=sc, unit_id=uid
        ).first()
        if rec:
            rec.qty = qty
            rec.fin = fin
            action = "update"
            updated += 1
        else:
            rec = OpsActualDaily(
                date=d, service_code=sc, unit_id=uid, qty=qty, fin=fin
            )
            db.session.add(rec)
            action = "create"
            created += 1
        db.session.add(
            AuditLog(
                entity="ops_actual_daily",
                entity_id=str(rec.id if rec.id else 0),
                action=action,
                actor="system",
                at=datetime.utcnow(),
                diff_json={},
            )
        )
    db.session.commit()
    return jsonify({"created": created, "updated": updated, "deactivated": 0})


@api_bp.post("/import/calendar-1404")
def import_calendar() -> Response:
    mode = request.args.get("mode", "dry-run")
    file = request.files.get("file")
    if not file:
        return jsonify({"errors": [{"row": 0, "errors": ["no file"]}]}), 400
    rows, errors = _parse_calendar(file)
    existing = {c.date_shamsi for c in CalendarDim.query.all()}
    file_dates = {r["date_shamsi"] for r in rows}
    created_list = sorted(list(file_dates - existing))
    updated_list = sorted(list(file_dates & existing))
    if mode == "dry-run":
        return jsonify({"errors": errors, "created": created_list, "updated": updated_list})
    if errors:
        return jsonify({"errors": errors}), 400
    created = updated = 0
    for row in rows:
        rec = CalendarDim.query.filter_by(date_shamsi=row["date_shamsi"]).first()
        data = dict(
            jalali_month=int(row["jalali_month"]),
            weekday_name=row["weekday_name"],
            is_friday=row.get("is_friday") == "1",
            is_thursday=row.get("is_thursday") == "1",
            is_official_holiday=row.get("is_official_holiday") == "1",
            is_summer_break=row.get("is_summer_break") == "1",
            weight_raw=float(row["weight_raw"]),
        )
        if rec:
            for k, v in data.items():
                setattr(rec, k, v)
            updated += 1
            action = "update"
        else:
            rec = CalendarDim(date_shamsi=row["date_shamsi"], **data)
            db.session.add(rec)
            created += 1
            action = "create"
        db.session.add(
            AuditLog(
                entity="calendar_dim",
                entity_id=str(rec.id if rec.id else 0),
                action=action,
                actor="system",
                at=datetime.utcnow(),
                diff_json={},
            )
        )
    db.session.commit()
    return jsonify({"created": created, "updated": updated})


@api_bp.post("/import/seasonality")
def import_seasonality() -> Response:
    mode = request.args.get("mode", "dry-run")
    file = request.files.get("file")
    if not file:
        return jsonify({"errors": [{"row": 0, "errors": ["no file"]}]}), 400
    rows, errors = _parse_seasonality(file)
    existing = {s.month for s in SeasonalityMonth.query.all()}
    file_months = {int(r["month"]) for r in rows if r.get("month") not in (None, "")}
    created_list = sorted(list(file_months - existing))
    updated_list = sorted(list(file_months & existing))
    if mode == "dry-run":
        return jsonify({"errors": errors, "created": created_list, "updated": updated_list})
    if errors:
        return jsonify({"errors": errors}), 400
    created = updated = 0
    for row in rows:
        month = int(row["month"])
        rec = SeasonalityMonth.query.filter_by(month=month).first()
        data = dict(
            actual_1403=float(row["actual_1403"]),
            season_weight=float(row["season_weight"]),
        )
        if rec:
            for k, v in data.items():
                setattr(rec, k, v)
            updated += 1
            action = "update"
        else:
            rec = SeasonalityMonth(month=month, **data)
            db.session.add(rec)
            created += 1
            action = "create"
        db.session.add(
            AuditLog(
                entity="seasonality_month",
                entity_id=str(rec.id if rec.id else 0),
                action=action,
                actor="system",
                at=datetime.utcnow(),
                diff_json={},
            )
        )
    db.session.commit()
    return jsonify({"created": created, "updated": updated})


@api_bp.get("/templates/services.csv")
def template_services() -> Response:
    output = "code,name,uom,base_qty,base_fin,is_active\n"
    return Response(output, mimetype="text/csv")


@api_bp.get("/templates/budget-annual.csv")
def template_budget() -> Response:
    output = "year,scenario,version,service_code,unit_id,annual_qty,annual_fin,currency\n"
    return Response(output, mimetype="text/csv")


@api_bp.get("/templates/ops-actual.csv")
def template_ops_actual() -> Response:
    output = "date_shamsi,date_gregorian,service_code,unit_id,qty,fin\n"
    return Response(output, mimetype="text/csv")


def _xlsx_response(wb: Workbook, filename: str) -> Response:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@api_bp.get("/templates/services.xlsx")
def template_services_xlsx() -> Response:
    wb = Workbook()
    ws = wb.active
    ws.title = "Services"
    ws.append(["code", "name", "uom", "base_qty", "base_fin", "is_active"])
    return _xlsx_response(wb, "services_template.xlsx")


@api_bp.get("/templates/budget-annual.xlsx")
def template_budget_xlsx() -> Response:
    wb = Workbook()
    ws = wb.active
    ws.title = "BudgetAnnual"
    ws.append([
        "year",
        "scenario",
        "version",
        "service_code",
        "unit_id",
        "annual_qty",
        "annual_fin",
        "currency",
        "notes",
    ])
    return _xlsx_response(wb, "budget_annual_template.xlsx")


@api_bp.get("/templates/ops-actual.xlsx")
def template_ops_actual_xlsx() -> Response:
    wb = Workbook()
    ws = wb.active
    ws.title = "OpsActualDaily"
    ws.append([
        "date_shamsi",
        "date_gregorian",
        "service_code",
        "unit_id",
        "qty",
        "fin",
    ])
    return _xlsx_response(wb, "ops_actual_template.xlsx")


@api_bp.get("/templates/calendar-1404.xlsx")
def template_calendar_xlsx() -> Response:
    wb = Workbook()
    ws = wb.active
    ws.title = "Calendar"
    ws.append(
        [
            "date_shamsi",
            "jalali_month",
            "weekday_name",
            "is_friday",
            "is_thursday",
            "is_official_holiday",
            "is_summer_break",
            "weight_raw",
        ]
    )
    ws["H2"] = "=IF(OR(D2=1,F2=1,G2=1),0,IF(E2=1,0.5,1))"
    return _xlsx_response(wb, "calendar_1404_template.xlsx")


@api_bp.get("/templates/seasonality.xlsx")
def template_seasonality_xlsx() -> Response:
    wb = Workbook()
    ws = wb.active
    ws.title = "SeasonalityMonth"
    ws.append(["month", "actual_1403", "season_weight"])
    for m in range(1, 13):
        row = m + 1
        ws.append([m, None, f"=B{row}/SUM($B$2:$B$13)"])
    return _xlsx_response(wb, "seasonality_template.xlsx")
